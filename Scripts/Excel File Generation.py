import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# File path of the original Excel file (Replace with actual file path)
input_file = r'path/to/input_file.xlsx'

# Directory where you want to save the Excel files (Replace with actual directory path)
output_directory = r'path/to/output_directory'

# List of columns to be treated as text
text_columns = [
    'Merchant ID',
    'Bank Routing Number',
    'Bank Account Number',
    'Contact Phone',
    'Company Representative Contact',
    'Primary Contact Phone',
    'Legal Entity ID',
    'VAP Submerchant IDs',
    'Sub-Merchant ID',
    'Company Representative SSN',  # Redacted as this is sensitive data
    'Business Postal Code',
    'Street Address 1'
]

# List of columns to be treated as numbers
number_columns = [
    'Max Transaction Amount',
    'Stake Percent',
    'Annual Sales Volume',
    'Business Location',
    'Bank Account'
]

# Read the Excel file into a DataFrame, specifying dtypes
dtype_spec = {col: str for col in text_columns}
dtype_spec.update({col: float for col in number_columns})
df = pd.read_excel(input_file, dtype=dtype_spec)

# Replace NaN values with empty strings
df = df.fillna('')

# Replace empty strings with 'n/a'
df = df.replace('', 'n/a')

# Convert remaining columns to text if not specified as number columns
for col in df.columns:
    if col not in number_columns:
        df[col] = df[col].astype(str)

# Group the data by the "Group Name" column (formerly "Chain Name")
grouped = df.groupby("Group Name")

# Function to set column widths
def set_column_widths(sheet, df):
    for col in df.columns:
        max_length = max(df[col].astype(str).map(len).max(), len(col))
        col_idx = df.columns.get_loc(col) + 1
        sheet.column_dimensions[sheet.cell(row=1, column=col_idx).column_letter].width = max_length + 2

# Iterate over each group and save it to a separate Excel file
for name, group in grouped:
    # Remove the "Group Name" column from the group
    group = group.drop(columns=["Group Name"])

    # Create a new workbook
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = name

    # Write the header to the new sheet
    for col_num, column_title in enumerate(group.columns, 1):
        new_sheet.cell(row=1, column=col_num, value=column_title)

    # Write the data to the new sheet starting from the second row
    for r_idx, row in enumerate(dataframe_to_rows(group, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            new_cell = new_sheet.cell(row=r_idx, column=c_idx, value=value)
            if group.columns[c_idx - 1] in text_columns:
                new_cell.number_format = '@'  # Set text format for specified columns
            elif group.columns[c_idx - 1] in number_columns:
                new_cell.number_format = '0'  # Set integer format for numeric columns
            else:
                new_cell.number_format = '@'  # Set default format as text

    # Set column widths to fit the contents
    set_column_widths(new_sheet, group)

    # Save the new workbook
    filename = os.path.join(output_directory, f"{name}.xlsx")
    new_workbook.save(filename)
